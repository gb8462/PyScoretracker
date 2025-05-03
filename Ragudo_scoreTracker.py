import tkinter as tk
from tkinter import ttk
import openpyxl, os

main = tk.Tk()
main.title('ScoreTracker')
main.geometry('750x400')
main.config(bg='#e4eaeb')

# ========== Functions ==========

def initialize_database():
    if not os.path.exists("student_scores.xlsx"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "ScoreTracker"
        sheet.append(["studentNames", "studentScores", "Result"])
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        workbook.save("student_scores.xlsx")

def Submit():
    student_name = name_entry.get()
    student_score = score_entry.get()

    if student_name.strip() == "" or student_score.strip() == "":
        print("Please enter both name and score.")
        return

    try:
        score = float(student_score)

        if score < 0 or score > 100:
            print("Score must be between 0 and 100.")
            return

        result = "Pass" if score >= 75 else "Fail"

        workbook = openpyxl.load_workbook("student_scores.xlsx")
        sheet = workbook.active

        while True:
            last_row_val = sheet.cell(row=sheet.max_row, column=1).value
            if last_row_val == "Average" or (last_row_val is None or str(last_row_val).strip() == ""):
                sheet.delete_rows(sheet.max_row, 1)
            else:
                break
            
        sheet.append([student_name, score, result])

        last_row = sheet.max_row

        scores = []
        for row in sheet.iter_rows(min_row=2, max_row=last_row, min_col=2, max_col=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    scores.append(cell.value)

        average_score = sum(scores) / len(scores) if scores else 0

        sheet.append(["", "", ""])
        avg_row = sheet.max_row + 1
        sheet.cell(row=avg_row, column=1).value = "Average"
        sheet.cell(row=avg_row, column=2).value = round(average_score, 2)

        workbook.save("student_scores.xlsx")

        print("Data saved!")

        name_entry.delete(0, tk.END)
        score_entry.delete(0, tk.END)

        load_data()

    except ValueError:
        print("Score must be a number.")
    except Exception as e:
        print("Error saving data:", e)

def load_data():
    try:
        workbook = openpyxl.load_workbook("student_scores.xlsx")
        sheet = workbook.active

        for row in tree.get_children():
            tree.delete(row)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, score, result = row
            if name is None and score is None and result is None:
                continue
            tree.insert("", tk.END, values=(name, score, result))

    except Exception as e:
        print("Error loading data:", e)

# ========== Frame1 (Form) ==========
frame1 = tk.Frame(main, bg='#e4eaeb', borderwidth=1, relief='flat')
frame1.pack(side='left', fill='both')

# ========== Title ==========
title = tk.Label(frame1, text='ScoreTracker', font=('Arial', 20), bg='#e4eaeb')
title.pack(pady=(20, 0))

# ========== Name Entry ==========
studentName = tk.Label(frame1, text='Student Name', font=('Arial', 12), bg='#e4eaeb')
studentName.pack(pady=(40, 2), padx=20)

name_entry = tk.Entry(frame1, width=20, font=('Arial', 11))
name_entry.pack(padx=20)

# ========== Score Entry ==========
score = tk.Label(frame1, text='Student Score', font=('Arial', 12), bg='#e4eaeb')
score.pack(pady=(10, 2), padx=20)

score_entry = tk.Entry(frame1, width=20, font=('Arial', 11))
score_entry.pack(padx=20)

# ========== Submit Button ==========
submit = tk.Button(frame1, text='Submit', command=Submit)
submit.pack(pady=20, padx=20)

# ========== Frame2 (Display Box) ==========
frame2 = tk.Frame(main,bg='#d3dadb', borderwidth=1, relief="sunken", highlightthickness=1)
frame2.pack(padx=10, pady=10, side='right', fill='both', expand=True)

tree = ttk.Treeview(frame2, columns=("Name", "Score", "Result"), show='headings')
tree.heading("Name", text="Name")
tree.heading("Score", text="Score")
tree.heading("Result", text="Result")

tree.column("Name", width=200)
tree.column("Score", width=100, anchor='center')
tree.column("Result", width=100, anchor='center')

tree.pack(fill='both', expand=True)

# ========== Init ==========
initialize_database()
load_data()
main.mainloop()
